import openpyxl, argparse, glob, csv, json, os, re, sys
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font

# Define the regular expression pattern to extract the desired information
#pattern = r"\](.+) \[(.+)\] {(.+)}<(.+)>\((.+?)\).+?(.*)"
pattern = r'\[.*?\]\s+(\d{4}-\d{2}-\d{2} \d{2}:\d{2}:\d{2}Z)\s+\[(.*?)\]\s+{(.*?)}<([^>]+)>\((.*?)\)\s+(.*)'

isShare=False
isFile=False

def parse_arguments():
    parser = argparse.ArgumentParser(description='Parse Snaffler log file(s) and save data to CSV/XLSX.')
    group = parser.add_mutually_exclusive_group(required=True)
    group.add_argument('-l', '--log_file', help='Path to the log file containing the Snaffler logs')
    group.add_argument('-j', '--json_file', help='Process the json file containing the Snaffler logs')
    group.add_argument('-d', '--directory', help='Process all files containing the Snaffler logs with the specified extension in the current directory')
    parser.add_argument('-x', '--file_extension', help='File extension to filter files when using -d option')
    parser.add_argument('-o', '--output_file', default='snaffler_logs', help='Output file name')

    # Check that at least one of -j, -l, or -d is present
    if not any(parser.parse_args().__dict__.values()):
        parser.error('At least one of the arguments -l/--log_file, -j/--json_file or -d/--directory is required.')

    return parser.parse_args()

def set_headers(sheet):
    headers = ['Timestamp', 'Resource Type', 'Triage Color Level', 'File Path', 'File Name','Match Content or File Type']
    for col_num, header in enumerate(headers, 1):
        col_letter = get_column_letter(col_num)
        cell = sheet[f"{col_letter}1"]
        cell.value = header
        cell.font = Font(bold=True)
    sheet.auto_filter.ref = sheet.dimensions
    
def sanitize_data(data):
    # Define a regular expression to match invalid characters
    invalid_char_pattern = re.compile(r'[\x00-\x08\x0B\x0C\x0E-\x1F]')

    sanitized_data = invalid_char_pattern.sub('_', data)

    return sanitized_data

def process_log_file(file_path, sheet, row_num):
    with open(file_path, 'r') as file:
        file_path=''
        file_name=''
        for line in file:
            matches = re.findall(pattern, line)
            
            if matches:
                timestamp=matches[0][0]
                triage_level=matches[0][2]
                res_type=matches[0][1]
                if res_type=='Share':
                    file_path=matches[0][3]
                    file_name=matches[0][3].split('\\')[-1]
                else:
                    file_path=matches[0][4]
                    file_name=matches[0][4].split('\\')[-1]
                line_content = matches[0][5]
                content = re.sub(r'\\r\\n', '', line_content)
                content = re.sub(r'\\r|\\n|\\t', '', content)
                content = re.sub(r'\\\s', ' ', content)
                content = re.sub("=", "'=",content)
                
                row_data = [timestamp, res_type, triage_level, file_path, file_name, content]

                # Find the last row with data in the sheet
                last_row = sheet.max_row

                # Start appending data from the next row
                row_num = last_row + 1
                
                for col_num, data in enumerate(row_data, 1):
                    col_letter = get_column_letter(col_num)
                    sanitized_data = sanitize_data(data)
                    sheet[f"{col_letter}{row_num}"] = sanitized_data

                row_num += 1
    return row_num

def process_log_files(log_files, sheet, row_num, file_extension=None):
    for log_file_path in log_files:
        if file_extension:
            print(f"Parsing file: {log_file_path}")
            row_num = process_log_file(log_file_path, sheet, row_num)
            print(f"Total number of rows parsed: {row_num - 1}")
        elif not file_extension:  # Added condition to process all files when -x is not provided
            print(f"Parsing file: {log_file_path}")
            row_num = process_log_file(log_file_path, sheet, row_num)
            print(f"Total number of rows parsed: {row_num - 1}")

def save_workbook(workbook, output_file_name):
    output_file = f"{output_file_name}"
    workbook.save(output_file)
    return output_file

def write_csv(data, filename):
    header = ['Timestamp', 'Triage Color Level', 'File Path', 'File Name','Match Content or File Type']

    with open(filename, 'w', newline='', encoding='utf-8') as csvfile:
        csv_writer = csv.writer(csvfile, delimiter=';')
        csv_writer.writerow(header)
        csv_writer.writerows(data)



def parse_json_to_csv(json_data):
    global isFile
    global isShare
    entries = json_data.get('entries', [])
    csv_rows = []

    for entry in entries:
        entry_level = entry.get('level', '')

        # Skip entries where "level" is "Info"
        if entry_level == 'Info':
            continue

        timestamp = entry.get('time', '')

        event_properties = entry.get('eventProperties', {})

        # Check if "eventProperties" is empty
        if not event_properties:

            entry_message = entry.get('message', '')

            # Define the regex pattern
            pattern = r'\[.*?\]\s+(\d{4}-\d{2}-\d{2} \d{2}:\d{2}:\d{2}Z)\s+\[(.*?)\]\s+{(.*?)}<([^>]+)>\((.*?)\)\s+(.*)'

            # Use re.search to find the match
            match = re.search(pattern, entry_message)

            if match:
                triage_color_level = match.group(3)
                file_path = match.group(5)
                file_name=file_path.split("\\")[-1]
                match_context = match.group(6)
                csv_rows.append([timestamp.strip(), triage_color_level.strip(),\
                                  file_path.strip(), file_name.strip(), match_context.strip()])

            else:
                print("No match found for empty eventProperties.")

        else:
            triage_color_level = get_triage_color_level(event_properties)

            file_result={}
            share_result={}
            if isFile:
                file_result = event_properties.get(triage_color_level, {}).get('FileResult', {})
            elif isShare:
                share_result = event_properties.get(triage_color_level, {}).get('ShareResult', {})
            
            file_info = file_result.get('FileInfo', {})
            file_path = file_info.get('FullName', '')
            share_path = share_result.get('SharePath','')

            res_path=file_path if isFile else share_path if isShare else ''

            file_name = file_info.get('Name', '')

            text_result = file_result.get('TextResult', {})
            match_context = text_result.get('MatchContext', 'N/A')
            if match_context == "":
                match_context="Found sensitive file type"
            elif match_context == 'N/A' and isShare:
                match_context="Found readable share for path: {}".format(share_path)

            csv_rows.append([timestamp, triage_color_level, res_path, file_name, match_context])

    return csv_rows

def get_triage_color_level(event_properties):
    global isShare
    global isFile
    for color_level in event_properties.keys():
        if 'FileResult' in event_properties[color_level]:
            isFile=True
            return color_level
        elif 'ShareResult' in event_properties[color_level]:
            isShare=True
            return color_level
    return 'N/A'

def print_banner():
    banner = '''
      __           _   _                         
     (_  ._   _. _|_ _|_ |  _   _| |   _   _   _ 
     __) | | (_|  |   |  | (/_ (_| |_ (_) (_| _> 
                                           _|                        
    '''
    print(banner)

def main():
    
    args = parse_arguments()

    # remove dot at the beginning of the extension if present
    if args.file_extension and args.file_extension.startswith('.'):
        args.file_extension = args.file_extension[1:]
    
    if args.json_file:
        args.output_file += '.csv'
    else:
        args.output_file += '.xlsx'

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    set_headers(sheet)

    row_num = 2

    # analyze the log file passed as input
    if args.log_file:
        log_files = [args.log_file]
        process_log_files(log_files, sheet, row_num)
        output_file = save_workbook(workbook, args.output_file)

    # analyze the json file passed as input
    elif args.json_file and not args.log_file:
        with open(args.json_file, 'r') as json_file:
            json_data = json.load(json_file)
        csv_data = parse_json_to_csv(json_data)
        write_csv(csv_data, args.output_file)
        output_file = args.output_file
    
    # check the passed input directory and analyze ONLY the files from the extension used, if present otherwise analyze all file present in the directory
    elif args.directory and os.path.isdir(args.directory):
        log_files = glob.glob(os.path.join(args.directory, f"*.{args.file_extension}" if args.file_extension else "*"))
        process_log_files(log_files, sheet, row_num, file_extension=args.file_extension)
        output_file = save_workbook(workbook, args.output_file)
    
    else:
        sys.exit(1)

    print_banner()
    print(f"Log data has been successfully extracted and saved to {output_file}.")

if __name__ == "__main__":
    main()